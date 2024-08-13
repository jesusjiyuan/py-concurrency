# coding:utf-8
import os
import openpyxl as xl

def write_excel_file(folder_path,result):
    result_path = os.path.join(folder_path, "otb-out.xlsx")
    print(result_path)
    if os.path.exists(result_path):
        print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** n')
        workbook = xl.load_workbook(result_path)
    else:
        print('***** excel不存在，创建excel ' + result_path + ' ***** n')
        workbook = xl.Workbook()
        workbook.save(result_path)
    #sheet = workbook.create_sheet("test",2)
    sheet = workbook.active
    headers = ["1", "2", "3","4"]
    sheet.append(headers)
    #result = [['1', 1, 1], ['2', 2, 2], ['3', 3, 3]]
    for data in result:
        sheet.append(data)
    workbook.save(result_path)
    print('***** 生成Excel文件 ' + result_path + ' ***** n')
