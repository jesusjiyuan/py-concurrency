from pulp import LpVariable, LpMaximize, LpProblem, value, LpStatus, LpInteger, LpMinimize, lpSum, LpContinuous,LpAffineExpression,isNumber
from cla import MyLpVariable, MyLpAffineExpression
import openpyxl as xl
import os


# 第一个参数为这个问题取名字，第二个参数表示求目标函数的最大值（LpMaximize）
prob = LpProblem('min', sense=LpMaximize)
# name为变量名， lowBound为下边界，None为没边界, cat约束变量的类型LpInteger为整型]

#------1
#x = [0] 
#for i in range(1, 8, 1):
#    x.append(LpVariable(name='x'+str(i), lowBound=0, upBound=None,cat=LpInteger))
#
#print(x)
## # 约束条件
#prob += ((329*x[1] + 149*x[2] + 99*x[3] + 99*x[4] + 229*x[5] + 149*x[6] + 99*x[7])-500000) 
#prob += ((329*x[1] + 149*x[2] + 99*x[3] + 99*x[4] + 229*x[5] + 149*x[6] + 99*x[7])-500000)/500000 <= 0.02

#------2
rp=[0,329,149,99,99,229,149,99]

x = [[0]]
for i in range(1, 8, 1):
    tmp = [0]
    for j in range(1,5,1):
        tmp.append(LpVariable(name='x'+str(i)+str(j), lowBound=0, upBound=None,cat=LpInteger))
    x.append(tmp)


print(x)
# # 约束条件
#( print(rp[i]*x[i][j] ) for i in range(1,len(x),1) for j in range(1,len(x[1]),1))
#prob += lpSum( print(rp[i]*x[i][j] ) for i in range(1,len(x),1) for j in range(1,len(x[1]),1))  - 3212000

prob += ((329*x[1][1] + 149*x[2][1] + 99*x[3][1] + 99*x[4][1] + 229*x[5][1] + 149*x[6][1] + 99*x[7][1]) 
+ (329*x[1][2] + 149*x[2][2] + 99*x[3][2] + 99*x[4][2] + 229*x[5][2] + 149*x[6][2] + 99*x[7][2]) 
+ (329*x[1][3] + 149*x[2][3] + 99*x[3][3] + 99*x[4][3] + 229*x[5][3] + 149*x[6][3] + 99*x[7][3])
+ (329*x[1][4] + 149*x[2][4] + 99*x[3][4] + 99*x[4][4] + 229*x[5][4] + 149*x[6][4] + 99*x[7][4])) - 3212000
prob += ((329*x[1][1] + 149*x[2][1] + 99*x[3][1] + 99*x[4][1] + 229*x[5][1] + 149*x[6][1] + 99*x[7][1]) - 500000)/500000 <= 0.02
prob += -((329*x[1][1] + 149*x[2][1] + 99*x[3][1] + 99*x[4][1] + 229*x[5][1] + 149*x[6][1] + 99*x[7][1]) - 500000)/500000 >= -0.02
prob += ((329*x[1][2] + 149*x[2][2] + 99*x[3][2] + 99*x[4][2] + 229*x[5][2] + 149*x[6][2] + 99*x[7][2]) - 845000)/845000 <= 0.02
prob += -((329*x[1][2] + 149*x[2][2] + 99*x[3][2] + 99*x[4][2] + 229*x[5][2] + 149*x[6][2] + 99*x[7][2]) - 845000)/845000 >= -0.02
prob += ((329*x[1][3] + 149*x[2][3] + 99*x[3][3] + 99*x[4][3] + 229*x[5][3] + 149*x[6][3] + 99*x[7][3]) - 532000)/532000 <= 0.02
prob += -((329*x[1][3] + 149*x[2][3] + 99*x[3][3] + 99*x[4][3] + 229*x[5][3] + 149*x[6][3] + 99*x[7][3]) - 532000)/532000 >= -0.02
prob += ((329*x[1][4] + 149*x[2][4] + 99*x[3][4] + 99*x[4][4] + 229*x[5][4] + 149*x[6][4] + 99*x[7][4]) - 1335000)/1335000 <= 0.02
prob += -((329*x[1][4] + 149*x[2][4] + 99*x[3][4] + 99*x[4][4] + 229*x[5][4] + 149*x[6][4] + 99*x[7][4]) - 1335000)/1335000 >= -0.02

prob += ((329*x[1][1] + 149*x[2][1] + 99*x[3][1] + 99*x[4][1] + 229*x[5][1] + 149*x[6][1] + 99*x[7][1]) 
+ (329*x[1][2] + 149*x[2][2] + 99*x[3][2] + 99*x[4][2] + 229*x[5][2] + 149*x[6][2] + 99*x[7][2]) 
+ (329*x[1][3] + 149*x[2][3] + 99*x[3][3] + 99*x[4][3] + 229*x[5][3] + 149*x[6][3] + 99*x[7][3]) - 1877000)/1877000 <= 0.01


prob += ((329*x[1][4] + 149*x[2][4] + 99*x[3][4] + 99*x[4][4] + 229*x[5][4] + 149*x[6][4] + 99*x[7][4]) - 1335000)/1335000 <= 0.01

prob += ((329*x[1][1] + 149*x[2][1] + 99*x[3][1] + 99*x[4][1] + 229*x[5][1] + 149*x[6][1] + 99*x[7][1]) 
+ (329*x[1][2] + 149*x[2][2] + 99*x[3][2] + 99*x[4][2] + 229*x[5][2] + 149*x[6][2] + 99*x[7][2]) 
+ (329*x[1][3] + 149*x[2][3] + 99*x[3][3] + 99*x[4][3] + 229*x[5][3] + 149*x[6][3] + 99*x[7][3]) 
+ (329*x[1][4] + 149*x[2][4] + 99*x[3][4] + 99*x[4][4] + 229*x[5][4] + 149*x[6][4] + 99*x[7][4]) - 3212000)/3212000 <= 0.001

prob += MyLpAffineExpression(x[1][1]*1) % 10 == 0

#prob += x[1][1] >= 100
for i in range(1,len(x),1):
    for j in range(1,len(x[1]),1):
        print(x[i][j])
        prob += x[i][j] >= 100
        prob += x[i][j] <= 1200
#(print(x[i][j])  for i in range(1,len(x),1) for j in range(1,len(x[1]),1))
prob += x[1][1] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[2][1] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[3][1] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[4][1] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[5][1] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[6][1] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[7][1] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])

prob += x[1][2] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[2][2] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[3][2] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[4][2] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[5][2] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[6][2] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[7][2] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])

prob += x[1][3] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[2][3] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[3][3] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[4][3] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[5][3] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[6][3] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[7][3] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])

prob += x[1][4] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[2][4] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[3][4] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[4][4] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[5][4] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[6][4] >= 0.2*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])
prob += x[7][4] >= 0.1*(x[1][1]+x[2][1]+x[3][1]+x[4][1]+x[5][1]+x[6][1]+x[7][1])

status = prob.solve()
print("求解状态:", LpStatus[prob.status])
print(f"目标函数的最大值z={value(prob.objective)}，此时目标函数的决策变量为:",
      {v.name: v.varValue for v in prob.variables()})




def write_excel_file(folder_path):
  result_path = os.path.join(folder_path, "jiyuan.xlsx")
  print(result_path)
  if os.path.exists(result_path):
    print('***** excel已存在，在表后添加数据 ' + result_path + ' ***** n')
    workbook = xl.load_workbook(result_path)
  else:
    print('***** excel不存在，创建excel ' + result_path + ' ***** n')
    workbook = xl.Workbook()
    workbook.save(result_path)
  sheet = workbook.create_sheet("test",2)
  headers = ["URL", "predict", "score"]
  sheet.append(headers)
  result = [['1', 1, 1], ['2', 2, 2], ['3', 3, 3]]
  for data in result:
    sheet.append(data)
  workbook.save(result_path)
  print('***** 生成Excel文件 ' + result_path + ' ***** n')

#if __name__ == '__main__':
#  write_excel_file("D:\DOC\docker-k8s\conf\算法")